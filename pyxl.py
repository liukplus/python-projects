from openpyxl import Workbook, load_workbook, Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
data = {
	"Joe": {
		"math": 65,
		"science": 78,
		"english": 98,
		"gym": 89
	},
	"Bill": {
		"math": 55,
		"science": 72,
		"english": 87,
		"gym": 95
	},
	"Tim": {
		"math": 100,
		"science": 45,
		"english": 75,
		"gym": 92
	},
	"Sally": {
		"math": 30,
		"science": 25,
		"english": 45,
		"gym": 100
	},
	"Jane": {
		"math": 100,
		"science": 100,
		"english": 100,
		"gym": 60
	}
}
wb = Workbook()
ws = wb.active
number_of_students = len(data)
number_of_subjects = len(data['Joe'])
ws.title = "Grades"
headings = ['Name'] + list(data['Joe'].keys())
ws.append(headings)

for student in data:
    grades = list(data[student].values())
    ws.append([student]+grades)

for col in range(2,number_of_subjects+2):
	char = get_column_letter(col)
	ws[char+str(number_of_students+2)] = f"=sum({char+str(2)}:{char+str(number_of_subjects+2)})/{len(data)}"
for col in range(1,number_of_subjects+2):
	ws[get_column_letter(col)+"1"].font = Font(bold=True,color='000000FF')
wb.save('Grades.xlsx')

